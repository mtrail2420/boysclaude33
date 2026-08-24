#!/usr/bin/env python3
"""
Run this once a year (March, after awards — or April, after the draft)
when Trail_Bish_Dynasty_Premium.xlsx has been updated.

Usage:
    python3 scripts/build_data.py path/to/Trail_Bish_Dynasty_Premium.xlsx

Regenerates data/players.json, data/man_status.json, and data/wildcards.json.

SCORING (v4, PFR-verified):
Reads real award/All-Pro/draft-class data from data/pfr_source/*.json (built
from Pro Football Reference exports). A player matched in that data is scored
from real selection counts. A player NOT matched falls back to parsing their
notes text for role language only (no awards guessed from prose).

Pro Bowl selections are NOTED in a player's generated note text but are never
part of the score — the owner's explicit call: Pro Bowl voting is considered
too inconsistent to weight fairly.

To refresh the PFR source maps themselves (e.g. after a new season of AP
All-Pro rosters or a new draft class becomes available), re-run the one-off
parsing steps used to build allpro_map.json / awards_map.json / draft_map.json
from new PFR workbook exports, and drop the updated JSON files into
data/pfr_source/. This script does not re-parse raw PFR .xls/.xlsx files
itself — it consumes the pre-built maps for speed and stability.
"""
import sys
import re
import json
import pathlib
import openpyxl
from collections import Counter

AWARD_WEIGHTS = {'MVP': 40, 'SB_MVP': 25, 'OPOY': 22, 'DPOY': 22, 'OROY': 12, 'DROY': 12}

# confirmed name aliases (nicknames PFR/awards data doesn't use)
NAME_ALIASES = {"RG III": "Robert Griffin III"}


def norm(n):
    n = re.sub(r"[.'’]", '', n or '')
    n = re.sub(r'\s+(Jr|Sr|II|III|IV|V)\.?$', '', n, flags=re.I)
    return re.sub(r'[^a-z]', '', n.lower())


def load_pfr_maps(data_dir):
    src = data_dir / "pfr_source"
    allpro = json.load(open(src / "allpro_map.json"))
    awards = json.load(open(src / "awards_map.json"))
    draft = json.load(open(src / "draft_map.json"))
    allpro_norm = {norm(k): (k, sorted(v)) for k, v in allpro.items()}
    draft_norm = {norm(k): (k, v) for k, v in draft.items()}
    awards_norm = {aw: {norm(k): (k, sorted(v)) for k, v in m.items()} for aw, m in awards.items()}
    return allpro_norm, draft_norm, awards_norm


def old_role_score(text):
    # role only -- Pro Bowl deliberately excluded from scoring. Broadened to catch
    # real career language beyond a narrow keyword set (e.g. "franchise QB", "elite
    # seasons") that earlier versions missed and under-scored real contributors for.
    if not text:
        return 0
    if re.search(r'did not stick|never played meaningful snaps|never established|failed to (stick|develop)|short[- ]lived NFL|never materialized|practice squad(?! member)', text, re.I):
        return 0
    if re.search(r'\belite\b|premier|franchise|productive|consistent|reliable|versatile|long[- ](term|serving)|leading|\bimpact\b|key contributor|start(er|ing|ed)?|contributed|playoff|champion|\d[,\d]*[- ]yard|double[- ]digit|over [\d,]+ career|career (yards|sacks|tackles|passing)|multiple.*(seasons|years)', text, re.I):
        return 5
    if re.search(r'\bdepth\b|backup|reserve|rotational|journeyman|\bbrief\b|limited role|spent time with (multiple|several)', text, re.I):
        return 2
    return 0


def tier_v4(s):
    if s >= 55: return "Legend"
    if s >= 28: return "Franchise"
    if s >= 12: return "High-End Starter"
    if s >= 5: return "Starter"
    if s >= 2: return "Contributor"
    return "Bust"


def score_player(name, existing_notes, allpro_norm, draft_norm, awards_norm):
    """Returns (score, tier, notes). existing_notes is the fallback source for
    unmatched (no PFR match at all) players' role scoring only. A player with a
    real major award or AP First-Team selection gets flat role credit rather than
    re-deriving role from their own (possibly already-regenerated) note text --
    that keeps re-runs of this script idempotent instead of losing information
    on every pass."""
    key = norm(NAME_ALIASES.get(name, name))
    pts = 0.0
    fact_bits = []
    pb_note = None
    has_award_or_allpro = False
    has_draft = False

    for aw, wt in AWARD_WEIGHTS.items():
        m = awards_norm[aw].get(key)
        if m:
            _, years = m
            pts += wt * len(years)
            fact_bits.append(f"{aw.replace('_', ' ')} ({', '.join(map(str, years))})")
            has_award_or_allpro = True

    ap = allpro_norm.get(key)
    if ap:
        _, years = ap
        pts += 15 * len(years)
        fact_bits.append(f"{len(years)}x AP First-Team All-Pro ({', '.join(map(str, years))})")
        has_award_or_allpro = True

    dr = draft_norm.get(key)
    role_pts = None
    if dr:
        _, d = dr
        pb = d.get('pb') or 0
        wav = d.get('wav') or 0
        if pb:
            pb_note = f"{pb}x Pro Bowl"  # noted, never scored
        role_pts = 5 if wav >= 18 else 2 if wav >= 5 else 0
        has_draft = True

    if role_pts is None and has_award_or_allpro:
        role_pts = 5  # a real major award/All-Pro selection already proves starter-caliber play
    if role_pts is None:
        role_pts = old_role_score(existing_notes)
    pts += role_pts

    score = round(pts, 1)
    tier = tier_v4(score)

    matched = has_award_or_allpro or has_draft
    if not matched:
        # no PFR match: score is still the role-fallback credit computed above,
        # notes stay as whatever was already there (nothing to regenerate)
        return score, tier, existing_notes

    note_parts = fact_bits[:4] if fact_bits else ["No AP All-Pro or major award found in the PFR record"]
    if pb_note:
        note_parts.append(pb_note)
    notes = "; ".join(note_parts) + ". (PFR-verified)"
    return score, tier, notes


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 build_data.py <path-to-workbook.xlsx>")
        sys.exit(1)

    xlsx_path = pathlib.Path(sys.argv[1])
    out_dir = pathlib.Path(__file__).parent.parent / "data"
    out_dir.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- PLAYERS ---
    ws = wb["PLAYERS"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r[0] == "RANK")
    players = []
    for r in rows[header_idx + 1:]:
        if r[0] is None:
            continue
        rank, name, pos, owner, year, score, tier, notes = r[:8]
        players.append({
            "rank": rank, "name": name, "position": pos, "owner": owner,
            "year": year, "notes": notes,
            # a blank score in the source workbook means "not yet graded" (e.g. this
            # year's rookies) -- that's an intentional signal, not missing data, and
            # a PFR match (a rookie's brand-new, all-zero draft-class row) must never
            # override it into a scored "Bust".
            "_originally_pending": score is None or str(score).strip() in ("", "\u2014", "-"),
        })

    allpro_norm, draft_norm, awards_norm = load_pfr_maps(out_dir)

    for p in players:
        if p.pop("_originally_pending"):
            p["score"] = "\u2014"
            p["tier"] = "Pending"
            continue
        s, t, n = score_player(p["name"], p["notes"], allpro_norm, draft_norm, awards_norm)
        p["score"] = f"{s:.1f}" if s is not None else "\u2014"
        p["tier"] = t if t else "Pending"
        if n is not None:
            p["notes"] = n

    players.sort(key=lambda p: (p["score"] == "\u2014", -(float(p["score"]) if p["score"] != "\u2014" else 0), p["rank"]))
    for i, p in enumerate(players, start=1):
        p["rank"] = i
    players.sort(key=lambda p: p["rank"])

    with open(out_dir / "players.json", "w") as f:
        json.dump(players, f, separators=(",", ":"))
    print(f"players.json — {len(players)} picks")

    # --- MAN STATUS ---
    ws2 = wb["MAN STATUS"]
    stats = []
    KEEP_RAW = {"Total Picks", "MVP Awards", "SB Wins", "All-Pro (total)", "SB MVP", "OPOY", "DPOY"}
    for r in ws2.iter_rows(values_only=True):
        if r[1] and r[1] in KEEP_RAW and r[2] is not None:
            stats.append({"label": r[1], "matt": r[2], "ryan": r[3], "edge": r[4]})

    def recompute(owner):
        ps = [p for p in players if p["owner"] == owner and p["score"] != "\u2014"]
        scores = [float(p["score"]) for p in ps]
        avg = round(sum(scores) / len(scores), 2) if scores else 0
        return avg, Counter(p["tier"] for p in ps)

    m_avg, m_c = recompute("Matt")
    r_avg, r_c = recompute("Ryan")

    def add(label, mval, rval):
        edge = "TIED" if mval == rval else ("MATT" if mval > rval else "RYAN")
        stats.append({"label": label, "matt": str(mval), "ryan": str(rval), "edge": edge})

    add("Avg Score", m_avg, r_avg)
    add("Franchise", m_c.get("Franchise", 0), r_c.get("Franchise", 0))
    add("High-End Starter", m_c.get("High-End Starter", 0), r_c.get("High-End Starter", 0))
    add("Starters", m_c.get("Starter", 0), r_c.get("Starter", 0))
    add("Contributors", m_c.get("Contributor", 0), r_c.get("Contributor", 0))
    add("Busts", m_c.get("Bust", 0), r_c.get("Bust", 0))

    with open(out_dir / "man_status.json", "w") as f:
        json.dump(stats, f, separators=(",", ":"))
    print(f"man_status.json — {len(stats)} rows")

    # --- WILDCARD BOYS ---
    ws3 = wb["WILDCARD BOYS"]
    rows3 = list(ws3.iter_rows(values_only=True))
    wc = []
    for r in rows3[1:]:
        if r[0] is None or r[0] == "YEAR":
            continue
        year, owner, player, pos, cat, outcome, meter, notes = r[:8]
        wc.append({
            "year": year, "owner": owner, "player": player, "position": pos,
            "category": cat, "outcome": outcome, "meter": meter, "notes": notes,
        })
    with open(out_dir / "wildcards.json", "w") as f:
        json.dump(wc, f, separators=(",", ":"))
    print(f"wildcards.json — {len(wc)} entries")

    print("\nDone. Commit the data/ folder and redeploy.")
    print("NOTE: verification.json (the Verify page log) is not touched by this script —")
    print("update it separately if the PFR source maps changed which players are matched.")


if __name__ == "__main__":
    main()
